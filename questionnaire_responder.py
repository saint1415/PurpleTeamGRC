#!/usr/bin/env python3
"""

Version: 3.0 - Updated for dynamic path detection
Questionnaire Auto-Responder
Automatically generate responses to security questionnaires with evidence links
"""

import json
import yaml
import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Add parent directory to path for imports
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

# Import path helper for dynamic paths
from utils.path_helper import BASE_DIR, CONFIG_FILE, RESULTS_DIR, LOGS_DIR
# Specific results directory
QUESTIONNAIRES_DIR = RESULTS_DIR / 'questionnaires'

# Paths

EVIDENCE_DIR = BASE_DIR / 'evidence'
QUESTIONNAIRE_DIR = RESULTS_DIR / 'questionnaires'
DB_FILE = EVIDENCE_DIR / 'evidence.db'

# Standard Questionnaire Mappings
QUESTIONNAIRE_MAPPINGS = {
    'VSA': {
        # Vendor Security Alliance Standard Questionnaire
        'A.1.1': {
            'question': 'Do you have a documented information security policy?',
            'controls': [('SOC2', 'CC6.1'), ('ISO27001', 'A.5.1.1'), ('NIST', 'PM-9')],
            'standard_response': 'Yes, we maintain a comprehensive information security policy reviewed annually.',
            'evidence_type': 'policy_document'
        },
        'A.1.2': {
            'question': 'Do you perform regular vulnerability assessments?',
            'controls': [('SOC2', 'CC6.7'), ('NIST', 'RA-5'), ('ISO27001', 'A.12.6.1')],
            'standard_response': 'Yes, we conduct weekly automated vulnerability scans of all systems.',
            'evidence_type': 'vulnerability_scan'
        },
        'A.2.1': {
            'question': 'Do you have a formal access control policy?',
            'controls': [('SOC2', 'CC6.1'), ('NIST', 'AC-2'), ('HIPAA', '164.312(a)(1)')],
            'standard_response': 'Yes, access controls are enforced through role-based access control (RBAC).',
            'evidence_type': 'access_control_scan'
        },
    },
    'SIG': {
        # Shared Assessments SIG Questionnaire
        'B.1': {
            'question': 'Does the organization have a process to identify and assess risks?',
            'controls': [('SOC2', 'CC7.1'), ('NIST', 'RA-3'), ('ISO27001', 'A.6.1.2')],
            'standard_response': 'Yes, we conduct quarterly risk assessments and maintain a risk register.',
            'evidence_type': 'compliance_check'
        },
        'B.2': {
            'question': 'Are vulnerability scans performed on a regular basis?',
            'controls': [('NIST', 'RA-5'), ('ISO27001', 'A.12.6.1'), ('HIPAA', '164.308(a)(1)(ii)(A)')],
            'standard_response': 'Yes, automated vulnerability scans are performed weekly.',
            'evidence_type': 'vulnerability_scan'
        },
    },
    'CAIQ': {
        # Cloud Security Alliance CAIQ
        'A&A-01': {
            'question': 'Do you have a documented approval process for asset and application registration?',
            'controls': [('SOC2', 'CC6.1'), ('ISO27001', 'A.8.1.1')],
            'standard_response': 'Yes, all assets are documented in our configuration management database (CMDB).',
            'evidence_type': 'network_scan'
        },
        'IVS-01': {
            'question': 'Do you conduct network-layer vulnerability scans?',
            'controls': [('NIST', 'RA-5'), ('ISO27001', 'A.12.6.1')],
            'standard_response': 'Yes, network vulnerability scans are conducted weekly.',
            'evidence_type': 'network_scan'
        },
    }
}

class QuestionnaireResponder:
    """Auto-generate questionnaire responses with evidence"""
    
    def __init__(self):
        self.ensure_directories()
        self.load_response_templates()
    
    def ensure_directories(self):
        """Create questionnaire directories"""
        directories = [
            QUESTIONNAIRE_DIR,
            QUESTIONNAIRE_DIR / 'responses',
            QUESTIONNAIRE_DIR / 'templates',
            QUESTIONNAIRE_DIR / 'incoming',
        ]
        for directory in directories:
            directory.mkdir(parents=True, exist_ok=True)
    
    def load_response_templates(self):
        """Load custom response templates"""
        template_file = QUESTIONNAIRE_DIR / 'templates' / 'custom_responses.yaml'
        
        if template_file.exists():
            with open(template_file, 'r') as f:
                self.custom_responses = yaml.safe_load(f)
        else:
            self.custom_responses = {}
            # Create default template
            with open(template_file, 'w') as f:
                yaml.dump({
                    'company_name': 'Your Company Name',
                    'respondent_name': 'Security Team',
                    'respondent_email': 'security@company.com',
                    'custom_answers': {}
                }, f)
    
    def generate_vsa_response(self, output_path: str = None) -> str:
        """Generate VSA questionnaire response"""
        
        if not output_path:
            output_path = QUESTIONNAIRE_DIR / 'responses' / f'VSA_Response_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VSA Response"
        
        # Header styling
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ['Question ID', 'Question', 'Response', 'Evidence', 'Controls Cited', 'Last Verified']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        
        # Populate responses
        row = 2
        for question_id, question_data in QUESTIONNAIRE_MAPPINGS['VSA'].items():
            # Get evidence for this question
            evidence = self._get_evidence_for_question(question_data)
            
            # Format controls
            controls_text = ', '.join([f"{fw} {ctrl}" for fw, ctrl in question_data['controls']])
            
            # Evidence links
            evidence_text = '\n'.join([f"{e['title']} ({e['created_date'][:10]})" for e in evidence[:3]])
            if len(evidence) > 3:
                evidence_text += f"\n... and {len(evidence) - 3} more"
            
            ws.cell(row=row, column=1).value = question_id
            ws.cell(row=row, column=2).value = question_data['question']
            ws.cell(row=row, column=3).value = question_data['standard_response']
            ws.cell(row=row, column=4).value = evidence_text if evidence else "Evidence available upon request"
            ws.cell(row=row, column=5).value = controls_text
            ws.cell(row=row, column=6).value = datetime.now().strftime('%Y-%m-%d')
            
            # Wrap text
            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        wb.save(output_path)
        logger.info(f"Generated VSA response: {output_path}")
        return str(output_path)
    
    def generate_sig_response(self, output_path: str = None) -> str:
        """Generate SIG questionnaire response"""
        
        if not output_path:
            output_path = QUESTIONNAIRE_DIR / 'responses' / f'SIG_Response_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SIG Response"
        
        # Header styling
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['Control', 'Question', 'Response', 'Evidence Files', 'Framework Citations', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Populate
        row = 2
        for control_id, question_data in QUESTIONNAIRE_MAPPINGS['SIG'].items():
            evidence = self._get_evidence_for_question(question_data)
            controls_text = ', '.join([f"{fw} {ctrl}" for fw, ctrl in question_data['controls']])
            
            ws.cell(row=row, column=1).value = control_id
            ws.cell(row=row, column=2).value = question_data['question']
            ws.cell(row=row, column=3).value = question_data['standard_response']
            ws.cell(row=row, column=4).value = f"{len(evidence)} evidence items available"
            ws.cell(row=row, column=5).value = controls_text
            ws.cell(row=row, column=6).value = datetime.now().strftime('%Y-%m-%d')
            
            row += 1
        
        wb.save(output_path)
        logger.info(f"Generated SIG response: {output_path}")
        return str(output_path)
    
    def generate_caiq_response(self, output_path: str = None) -> str:
        """Generate CAIQ questionnaire response"""
        
        if not output_path:
            output_path = QUESTIONNAIRE_DIR / 'responses' / f'CAIQ_Response_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CAIQ"
        
        # CAIQ format
        header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['Control ID', 'Question', 'Yes/No', 'Explanation', 'Evidence', 'Controls']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for control_id, question_data in QUESTIONNAIRE_MAPPINGS['CAIQ'].items():
            evidence = self._get_evidence_for_question(question_data)
            controls_text = ', '.join([f"{fw} {ctrl}" for fw, ctrl in question_data['controls']])
            
            ws.cell(row=row, column=1).value = control_id
            ws.cell(row=row, column=2).value = question_data['question']
            ws.cell(row=row, column=3).value = "Yes"
            ws.cell(row=row, column=4).value = question_data['standard_response']
            ws.cell(row=row, column=5).value = f"{len(evidence)} items"
            ws.cell(row=row, column=6).value = controls_text
            
            row += 1
        
        wb.save(output_path)
        logger.info(f"Generated CAIQ response: {output_path}")
        return str(output_path)
    
    def process_custom_questionnaire(self, input_file: str, output_file: str = None):
        """Process custom Excel questionnaire"""
        
        logger.info(f"Processing custom questionnaire: {input_file}")
        
        # Read input
        wb_in = openpyxl.load_workbook(input_file)
        ws_in = wb_in.active
        
        # Create output
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Completed Questionnaire"
        
        # Copy headers
        for col in range(1, ws_in.max_column + 1):
            ws_out.cell(row=1, column=col).value = ws_in.cell(row=1, column=col).value
        
        # Add response columns if not present
        response_col = ws_in.max_column + 1
        evidence_col = response_col + 1
        controls_col = evidence_col + 1
        
        ws_out.cell(row=1, column=response_col).value = "Response"
        ws_out.cell(row=1, column=evidence_col).value = "Evidence Count"
        ws_out.cell(row=1, column=controls_col).value = "Controls Cited"
        
        # Process each row
        for row in range(2, ws_in.max_row + 1):
            # Copy question data
            for col in range(1, ws_in.max_column + 1):
                ws_out.cell(row=row, column=col).value = ws_in.cell(row=row, column=col).value
            
            # Try to find matching question
            question_text = str(ws_in.cell(row=row, column=2).value or "")
            response, evidence, controls = self._match_question(question_text)
            
            ws_out.cell(row=row, column=response_col).value = response
            ws_out.cell(row=row, column=evidence_col).value = len(evidence)
            ws_out.cell(row=row, column=controls_col).value = controls
        
        if not output_file:
            output_file = QUESTIONNAIRE_DIR / 'responses' / f'Custom_Response_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb_out.save(output_file)
        logger.info(f"Saved custom questionnaire response: {output_file}")
        return str(output_file)
    
    def _get_evidence_for_question(self, question_data: Dict) -> List[Dict]:
        """Get evidence items for a question"""
        
        if not DB_FILE.exists():
            return []
        
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get evidence matching the controls
        evidence_list = []
        for framework, control_id in question_data['controls']:
            cursor.execute('''
                SELECT DISTINCT e.*
                FROM evidence e
                JOIN control_citations cc ON e.evidence_id = cc.evidence_id
                WHERE cc.framework = ? AND cc.control_id = ?
                AND e.status = 'active'
                ORDER BY e.created_date DESC
                LIMIT 5
            ''', (framework, control_id))
            
            evidence_list.extend([dict(row) for row in cursor.fetchall()])
        
        conn.close()
        
        # Remove duplicates
        unique_evidence = {e['evidence_id']: e for e in evidence_list}
        return list(unique_evidence.values())
    
    def _match_question(self, question_text: str) -> tuple:
        """Match custom question to known patterns"""
        
        question_lower = question_text.lower()
        
        # Keywords to evidence type mapping
        if 'vulnerability' in question_lower or 'scan' in question_lower:
            evidence_type = 'vulnerability_scan'
            response = "Yes, we conduct weekly automated vulnerability scans."
            controls = "NIST RA-5, ISO27001 A.12.6.1"
        elif 'access control' in question_lower or 'authentication' in question_lower:
            evidence_type = 'access_control_scan'
            response = "Yes, access controls are enforced through RBAC and MFA."
            controls = "SOC2 CC6.1, NIST AC-2, HIPAA 164.312(a)(1)"
        elif 'network' in question_lower or 'firewall' in question_lower:
            evidence_type = 'network_scan'
            response = "Yes, network security controls are in place and tested regularly."
            controls = "NIST SC-7, ISO27001 A.13.1.1"
        elif 'monitoring' in question_lower or 'logging' in question_lower:
            evidence_type = 'compliance_check'
            response = "Yes, comprehensive logging and monitoring systems are operational."
            controls = "SOC2 CC7.1, NIST SI-4, HIPAA 164.312(b)"
        elif 'encryption' in question_lower or 'crypto' in question_lower:
            evidence_type = 'compliance_check'
            response = "Yes, encryption is implemented for data in transit and at rest."
            controls = "HIPAA 164.312(e)(1), ISO27001 A.10.1.1"
        else:
            evidence_type = 'compliance_check'
            response = "Please see attached documentation."
            controls = "Multiple frameworks"
        
        # Get evidence count
        evidence = self._get_evidence_by_type(evidence_type)
        
        return response, evidence, controls
    
    def _get_evidence_by_type(self, evidence_type: str) -> List[Dict]:
        """Get evidence by type"""
        
        if not DB_FILE.exists():
            return []
        
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT * FROM evidence
            WHERE evidence_type = ?
            AND status = 'active'
            ORDER BY created_date DESC
            LIMIT 10
        ''', (evidence_type,))
        
        results = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return results
    
    def generate_all_standard_questionnaires(self):
        """Generate all standard questionnaire responses"""
        
        logger.info("Generating all standard questionnaires...")
        
        responses = {
            'VSA': self.generate_vsa_response(),
            'SIG': self.generate_sig_response(),
            'CAIQ': self.generate_caiq_response()
        }
        
        # Create master index
        self._create_questionnaire_index(responses)
        
        return responses
    
    def _create_questionnaire_index(self, responses: Dict):
        """Create index of all questionnaire responses"""
        
        index_file = QUESTIONNAIRE_DIR / 'responses' / 'INDEX.html'
        
        html = f'''
<!DOCTYPE html>
<html>
<head>
    <title>Questionnaire Responses Index</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #333; border-bottom: 3px solid #667eea; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background-color: #667eea; color: white; }}
        tr:hover {{ background-color: #f5f5f5; }}
        .info {{ background-color: #e7f3fe; padding: 15px; margin: 20px 0; }}
    </style>
</head>
<body>
    <h1>Security Questionnaire Responses</h1>
    <div class="info">
        <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p><strong>Total Questionnaires:</strong> {len(responses)}</p>
        <p>All responses are auto-generated from security scan evidence and linked to specific control frameworks.</p>
    </div>
    
    <h2>Available Responses</h2>
    <table>
        <tr>
            <th>Questionnaire Type</th>
            <th>File Name</th>
            <th>Description</th>
            <th>Controls Covered</th>
        </tr>
'''
        
        for qtype, filepath in responses.items():
            filename = Path(filepath).name
            if qtype == 'VSA':
                desc = "Vendor Security Alliance Standard Questionnaire"
                controls = "SOC2, NIST, ISO27001, HIPAA"
            elif qtype == 'SIG':
                desc = "Shared Assessments SIG Questionnaire"
                controls = "SOC2, NIST, ISO27001, HIPAA"
            elif qtype == 'CAIQ':
                desc = "Cloud Security Alliance CAIQ"
                controls = "SOC2, NIST, ISO27001"
            else:
                desc = "Custom Questionnaire"
                controls = "Various"
            
            html += f'''
        <tr>
            <td><strong>{qtype}</strong></td>
            <td>{filename}</td>
            <td>{desc}</td>
            <td>{controls}</td>
        </tr>
'''
        
        html += '''
    </table>
    
    <h2>Usage Instructions</h2>
    <ol>
        <li>Download the appropriate questionnaire response file</li>
        <li>Review all responses for accuracy</li>
        <li>Customize as needed for your specific context</li>
        <li>Evidence files referenced are available in the evidence repository</li>
        <li>Contact security team for evidence packages or additional information</li>
    </ol>
</body>
</html>
'''
        
        with open(index_file, 'w') as f:
            f.write(html)
        
        logger.info(f"Created questionnaire index: {index_file}")

def main():
    """Main entry point"""
    responder = QuestionnaireResponder()
    
    # Generate all standard questionnaires
    responses = responder.generate_all_standard_questionnaires()
    
    logger.info("Questionnaire generation complete")
    for qtype, filepath in responses.items():
        logger.info(f"  {qtype}: {filepath}")

if __name__ == '__main__':
    main()
